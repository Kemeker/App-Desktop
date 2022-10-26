import openpyxl
import pandas as pd
from tkinter import ttk
import tkinter as tk
import datetime as dt
import os
import tkinter.messagebox as msg
from funçao_banco import *

lista_voltagem = ["110v", "220v", "360v", "110v~220v"]
lista_ferramentas = carrega_lista('Ferramentas')
lista_tecnicos = carrega_lista('Tecnicos')
lista_turnos = ["Manha", "Tarde", "Noite"]

'''-----------------------------------------------INTERFACE CADASTRO DO TECNICO---------------------------------------'''
def janela_cadastro_tecnicos():
    janela_cadastro = tk.Toplevel()
    janela_cadastro.geometry("800x400")
    janela_cadastro.title("Cadastrar Técnico")
    '''------------------------------------------------ADICIONAR INF NO BD------------------------------------------------'''
    def novo_tecnico():
        global lista_tecnicos
        BD: openpyxl.Workbook = abrir_BD()
        sh = BD.active
        colI = 0
        colF = 0
        novaLinha = 0
        for col in range(1, 100, 1):
            if sh.cell(row=1, column=col).value == 'Tecnicos':
                colI = col
                break
        for col in range(colI, 100, 1):
            if sh.cell(row=1, column=col + 1).value == None:
                colF = col
                break
        for lin in range(1, 1_000_000, 1):
            if sh.cell(row=lin, column=colI).value == None:
                novaLinha = lin
                break
        contador = 0
        infomacoes = [entrar_nome.get(), entrar_cpf.get(), entrar_turno.get(), entrar_radio.get(), entrar_equipe.get()]
        for col in range(colI, colF + 1, 1):
            sh.cell(row=novaLinha, column=col).value = infomacoes[contador]
            contador +=1
        fechar_BD(BD)
        lista_tecnicos = carrega_lista('Tecnicos')
        janela_cadastro.destroy()
        msg.showinfo('Cadastro de Tecnicos', 'Tecnico Cadastrado!')
    
       

    '''-------------------------------------------------Cadastro do Tecnico'''''''''''''''''''''''''''''''''''''''''''''''
    label_nome = tk.Label(janela_cadastro, text='Nome do Técnico:')
    label_nome.grid(row=0, column=0, padx=0, pady=0, columnspan=1)
    entrar_nome = tk.Entry(janela_cadastro)
    entrar_nome.grid(row=0, column=1, padx=0, pady=0, columnspan=1)

    '''---------------------------------------------------cadastro do cpf-------------------------------------------------'''
    label_cpf = tk.Label(janela_cadastro, text='CPF:')
    label_cpf.grid(row=2, column=0, padx=0, pady=0, columnspan=1)
    entrar_cpf = tk.Entry(janela_cadastro)
    entrar_cpf.grid(row=2, column=1, padx=0, pady=0, columnspan=1)

    label_radio = tk.Label(janela_cadastro, text='Nª RADIO:')
    label_radio.grid(row=3, column=0, padx=0, pady=0, columnspan=1)
    entrar_radio = tk.Entry(janela_cadastro)
    entrar_radio.grid(row=3, column=1, padx=0, pady=0, columnspan=1)

    label_turno = tk.Label(janela_cadastro, text='Turno')
    label_turno.grid(row=4, column=0, padx=0, pady=0, columnspan=1)
    entrar_turno = ttk.Combobox(janela_cadastro, values=lista_turnos)
    entrar_turno.grid(row=4, column=1, padx=0, pady=0, columnspan=1)

    label_nome_equipe = tk.Label(janela_cadastro, text='Nome da Equipe:')
    label_nome_equipe.grid(row=5, column=0, padx=0, pady=0, columnspan=1)
    entrar_equipe = tk.Entry(janela_cadastro)
    entrar_equipe.grid(row=5, column=1, padx=0, pady=0, columnspan=1)

    botao_cadastrar_tecnico = tk.Button(janela_cadastro, text='Cadastrar', command=novo_tecnico)
    botao_cadastrar_tecnico.grid(row=6, column=3, padx=0, pady=0, columnspan=1)
'''-----------------------------------------------INTERFACE PARA CADASTRO DE FERRAMENTAS-----------------------------'''
def interface_cadastro_ferramentas():
    janela_cadastro_ferramentas = tk.Toplevel()
    janela_cadastro_ferramentas.geometry("800x400")
    janela_cadastro_ferramentas.title("Cadastrar Ferramenta")

    def nova_ferramenta():
        BD: openpyxl.Workbook = abrir_BD()
        sh = BD.active
        colI = 0
        colF = 0
        novaLinha = 0
        for col in range(1, 100, 1):
            if sh.cell(row=1, column=col).value == 'Ferramentas':
                colI = col
                break
        for col in range(colI, 100, 1):
            if sh.cell(row=1, column=col + 1).value == None:
                colF = col
                break
        for lin in range(1, 1_000_000, 1):
            if sh.cell(row=lin, column=colI).value == None:
                novaLinha = lin
                break
        contador = 0
        infomacoes = [entry_descricao.get() ]
        for col in range(colI, colF + 1, 1):
            sh.cell(row=novaLinha, column=col).value = infomacoes[contador]
            contador +=1
        fechar_BD(BD)
        janela_cadastro_ferramentas.destroy()
        msg.showinfo('Cadastrar', 'Ferramenta Cadastrada!')

    

    
    label_descricao = tk.Label(janela_cadastro_ferramentas, text="Descrição da ferramenta")
    label_descricao.grid(row=1, column=0, padx=10, pady=10, sticky='nswe', columnspan=4)

    entry_descricao = tk.Entry(janela_cadastro_ferramentas)
    entry_descricao.grid(row=2, column=0, padx=10, pady=10, sticky='nswe', columnspan=4)

    label_tipo_unidade = tk.Label(janela_cadastro_ferramentas, text="Código da ferramenta")
    label_tipo_unidade.grid(row=3, column=0, padx=10, pady=10, sticky='nswe', columnspan=2)

    combobox_selecionar_tipo = ttk.Combobox(janela_cadastro_ferramentas, values=lista_tipos)
    combobox_selecionar_tipo.grid(row=3, column=2, padx=10, pady=10, sticky='nswe', columnspan=2)

    label_quant = tk.Label(janela_cadastro_ferramentas, text="Quantia em estoque")
    label_quant.grid(row=4, column=0, padx=10, pady=10, sticky='nswe', columnspan=2)

    entry_quant = tk.Entry(janela_cadastro_ferramentas)
    entry_quant.grid(row=4, column=2, padx=10, pady=10, sticky='nswe', columnspan=2)

    botao_criar_codigo = tk.Button(janela_cadastro_ferramentas, text="Cadastrar", command=nova_ferramenta)
    botao_criar_codigo.grid(row=5, column=0, padx=10, pady=10, sticky='nswe', columnspan=4)
'''-----------------------------------------------INTERFACE PARA FAZER SOLICITAÇAO DAS FERRAMENTAS--------------------------------'''
def janela_solicitacoes():
    '''funçao para solicitar'''
    
    janela = tk.Toplevel()
    janela.geometry("800x400")
    janela.title("Solicitaçao de Ferramentas")

    def nova_solicitacao():
        global lista_ferramentas
        BD: openpyxl.Workbook = abrir_BD()
        sh = BD.active
        colI = 0
        colF = 0
        novaLinha = 0
        for col in range(1, 100, 1):
            if sh.cell(row=1, column=col).value == 'Nome':
                colI = col
                break
        for col in range(colI, 100, 1):
            if sh.cell(row=1, column=col + 1).value == None:
                colF = col
                break
        for lin in range(1, 1_000_000, 1):
            if sh.cell(row=lin, column=colI).value == None:
                novaLinha = lin
                break
        contador = 0
        infomacoes = [entry_tecnico.get(), ferramentas.get(), selecionar_voltagem.get(), entry_unidade.get(), entry_data_saida.get(), entry_data_entrada.get()]
        for col in range(colI, colF + 1, 1):
            sh.cell(row=novaLinha, column=col).value = infomacoes[contador]
            contador +=1
        fechar_BD(BD)
        lista_ferramentas = carrega_lista('Ferramentas')
        janela.destroy()
        msg.showinfo('Solicitar', 'Solicitacao Concluida!')

    label_nome_tecnico = tk.Label(janela, text="Técnico:")
    label_nome_tecnico.grid(row=0, column=0, padx=10, pady=10)
    entry_tecnico = ttk.Combobox(janela, values=lista_tecnicos)
    entry_tecnico.grid(row=0, column=1, padx=0, pady=0)

    '''------------------------------------------------Caixa para digitar a ferramenta----------------------------------'''
    label_nome_ferramenta = tk.Label(janela, text="Ferramenta:")
    label_nome_ferramenta.grid(row=1, column=0, padx=5, pady=2)
    ferramentas = ttk.Combobox(janela, values=lista_ferramentas)
    ferramentas.grid(row=1, column=1, padx=5, pady=5)

    '''----------------------------------Caixa para selecionar a voltagem da ferramenta---------------------------------'''
    label_selecionar_voltagem = tk.Label(janela, text="Voltagem:")
    label_selecionar_voltagem.grid(row=1, column=2, padx=5, pady=5)
    selecionar_voltagem = ttk.Combobox(janela, values=lista_voltagem)
    selecionar_voltagem.grid(row=1, column=3, padx=5, pady=5)

    '''---------------------------------------Caixa para digitar a unidade de ferramentas-------------------------------'''
    label_unidade = tk.Label(janela, text="Unidade:")
    label_unidade.grid(row=1, column=4, padx=5, pady=5, )
    entry_unidade = tk.Entry(janela)
    entry_unidade.grid(row=1, column=5, padx=5, pady=5, )
    '''------------------------------------------------------------------------------------------------------------------'''
    '''-------------------------------------------Data de entrada e saida das ferramentas--------------------------------'''
    label_data_saida = tk.Label(janela, text="Data de saida:")
    label_data_saida.grid(row=2, column=0, padx=10, pady=10, columnspan=1)
    entry_data_saida = tk.Entry(janela)
    entry_data_saida.grid(row=2, column=1, padx=10, pady=10, columnspan=1)
    label_data_entrada = tk.Label(janela, text="Data de entrada:")
    label_data_entrada.grid(row=2, column=2, padx=10, pady=10, columnspan=1)
    entry_data_entrada = tk.Entry(janela)
    entry_data_entrada.grid(row=2, column=3, padx=10, pady=10, columnspan=1)
    '''------------------------------------------------------------------------------------------------------------------'''
    '''--------------------------------------Botao para cadastrar a solicitaçao--------------------------------------'''
    botao_cadastrar_solicitacao = ttk.Button(janela, text="Cadastrar Solicitaçao", command=nova_solicitacao)
    botao_cadastrar_solicitacao.grid(row=4, column=1, sticky=tk.E, padx=5, pady=5)

    '''-----------------------------------------------INTERFACE PRINCIPAL DA APLICAÇAO-------------------------------------'''


'''-----------------------------------------------INTERFACE PRINCIPAL DO SISTEMA-----------------------------------------------'''
aplication = tk.Tk()
aplication.geometry("900x600")
aplication.title("DBV Softwares & Sistemas")

nome_cadastro = tk.Label(aplication, text='CADASTRAR FERRAMENTAS')
nome_cadastro.grid(row=0, column=6, padx=0, pady=0, columnspan=1)

nome_tecnico = tk.Label(aplication, text='CADASTRAR TECNICO')
nome_tecnico.grid(row=0, column=4, padx=0, pady=0, columnspan=1)

nome_solicitacao = tk.Label(aplication, text='SOLICITAR FERRAMENTA')
nome_solicitacao.grid(row=0, column=1, padx=0, pady=0, columnspan=1)

'''-------------------------------------------------Botoes para direcionar a paginas------------------------------------'''
imagem_solicitar_ferramenta = tk.PhotoImage(file=r'App\imagens\Solicitar ferramenta.png')
botao_solicitar_ferramentas = ttk.Button(aplication, image=imagem_solicitar_ferramenta, command=janela_solicitacoes)
botao_solicitar_ferramentas.place()
botao_solicitar_ferramentas.grid(row=2, column=1, padx=0, pady=0, columnspan=1)

imagem_cadastro_tecnico = tk.PhotoImage(file=r'App\imagens\tecnico.png')
botao_cadastrar_tecnico = ttk.Button(aplication, image=imagem_cadastro_tecnico, command=janela_cadastro_tecnicos)
botao_cadastrar_tecnico.place()
botao_cadastrar_tecnico.grid(row=2, column=4, padx=0, pady=0, columnspan=1)

imagem_cadastro_ferramenta = tk.PhotoImage(file=r'App\imagens\cadastrar ferramenta.png')
botao_cadastrar_ferramenta = ttk.Button(aplication, image=imagem_cadastro_ferramenta, command=interface_cadastro_ferramentas)
                                        
botao_cadastrar_ferramenta.place()
botao_cadastrar_ferramenta.grid(row=2, column=6, padx=0, pady=0, columnspan=1)

'''--------------------------------------------Photo Image-----------------------------------------'''

aplication.mainloop()













