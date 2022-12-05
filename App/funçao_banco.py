import openpyxl

_caminho = r'App\banco_dados.xlsx'
def abrir_BD()-> openpyxl.Workbook:
    '''Retorna em uma variavel o banco de dados aberto.'''
    BD = openpyxl.load_workbook(_caminho)
    return BD
def fechar_BD(BD: openpyxl.Workbook):
    BD.save(_caminho)
    BD.close()
def carrega_lista(Ferramentas: str) -> list:
    BD = abrir_BD()
    pergunta = False
    lista = []
    for col in range(1,1000,1):
        if pergunta == True:
            break
        if BD.active.cell(row=1,column=col).value == Ferramentas:
            for lin in range(2,1_000_000,1):
                if BD.active.cell(row=lin,column=col).value != None:
                    lista.append(BD.active.cell(row=lin,column=col).value)
                else:
                    pergunta = True
                    break
    fechar_BD(BD)
    return lista
def carregar_tabela(Tecnicos: str) -> list:
    '''encontra a primeira coluna e extrai a tabela inteira'''
    BD = abrir_BD()
    pergunta = False
    tabela = []
    for colI in range(1, 1000, 1):
        if pergunta == True:
            break
        if BD.active.cell(row=1, column=colI).value == Tecnicos:
            for lin in range(2,1000,1):
                if BD.active.cell(row=lin, column=colI).value == None:
                    pergunta = True
                    break
                lista= []
                for col in range(colI,1000,1):
                    if BD.active.cell(row=lin, column=col).value != None:
                        lista.append(BD.active.cell(row=lin, column=col).value)
                    else:
                        break
                tabela.append(lista)
    fechar_BD(BD)
    return tabela





